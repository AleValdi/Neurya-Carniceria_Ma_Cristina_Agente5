"""Generador de reporte Excel para demo del Agente5.

Genera un Excel con 4 hojas:
- Estado de Cuenta: cada linea del estado de cuenta con su resultado
- Resumen por Dia: conteos y montos por dia
- Comparacion Produccion: folios creados vs produccion
- Polizas: comparacion linea por linea de polizas contables
"""

from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import AccionLinea, ResultadoLinea, ResultadoProceso


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

_FILL_HEADER = PatternFill('solid', fgColor='1F4E79')

_FILL_INSERT = PatternFill('solid', fgColor='C6EFCE')       # Verde
_FILL_CONCILIAR = PatternFill('solid', fgColor='BDD7EE')    # Azul claro
_FILL_OMITIR = PatternFill('solid', fgColor='D9D9D9')       # Gris
_FILL_SIN_PROCESAR = PatternFill('solid', fgColor='FFF2CC') # Amarillo
_FILL_ERROR = PatternFill('solid', fgColor='FFC7CE')        # Rojo
_FILL_DESCONOCIDO = PatternFill('solid', fgColor='FFCCCC')  # Naranja/rosa

_FILL_SI = PatternFill('solid', fgColor='C6EFCE')
_FILL_NO = PatternFill('solid', fgColor='FFC7CE')
_FILL_NA = PatternFill('solid', fgColor='D9D9D9')
_FILL_STRIPE = PatternFill('solid', fgColor='F2F2F2')

_FONT_HEADER = Font(bold=True, color='FFFFFF', size=11)
_FONT_NORMAL = Font(size=10)

_BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

_ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
_ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

_FMT_MONEY = '#,##0.00'

_ACCION_FILL = {
    AccionLinea.INSERT: _FILL_INSERT,
    AccionLinea.CONCILIAR: _FILL_CONCILIAR,
    AccionLinea.OMITIR: _FILL_OMITIR,
    AccionLinea.SIN_PROCESAR: _FILL_SIN_PROCESAR,
    AccionLinea.ERROR: _FILL_ERROR,
    AccionLinea.DESCONOCIDO: _FILL_DESCONOCIDO,
}


# ---------------------------------------------------------------------------
# Funciones de query (standalone, sin pytest)
# ---------------------------------------------------------------------------

def _leer_movimiento(cursor, folio: int) -> Optional[dict]:
    """Lee un movimiento de SAVCheqPM por folio."""
    cursor.execute("""
        SELECT Folio, Banco, Cuenta, Age, Mes, Dia, Tipo,
               Ingreso, Egreso, Concepto, Clase, FPago, TipoEgreso,
               Conciliada, Paridad, ParidadDOF, Moneda,
               Cia, Fuente, Oficina, CuentaOficina,
               TipoPoliza, NumPoliza, Capturo, Sucursal
        FROM SAVCheqPM
        WHERE Folio = ?
    """, (folio,))
    row = cursor.fetchone()
    if not row:
        return None
    cols = [desc[0] for desc in cursor.description]
    return dict(zip(cols, row))


def _leer_poliza(cursor, folio: int) -> List[dict]:
    """Lee lineas de poliza de SAVPoliza por DocFolio."""
    cursor.execute("""
        SELECT Poliza, Movimiento, Cuenta, SubCuenta,
               TipoCA, Cargo, Abono, Concepto,
               DocTipo, TipoPoliza, DocFolio
        FROM SAVPoliza
        WHERE Fuente = 'SAV7-CHEQUES' AND DocFolio = ?
        ORDER BY Poliza, Movimiento
    """, (folio,))
    cols = [desc[0] for desc in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _buscar_produccion(
    cursor, clase: str, egreso: float, age: int, mes: int,
) -> Optional[int]:
    """Busca folio en produccion por Clase + Egreso exacto."""
    cursor.execute("""
        SELECT TOP 1 Folio
        FROM DBSAV71.dbo.SAVCheqPM
        WHERE RTRIM(Clase) = ? AND Egreso = ? AND Age = ? AND Mes = ?
        ORDER BY Folio DESC
    """, (clase, egreso, age, mes))
    row = cursor.fetchone()
    return row[0] if row else None


def _leer_movimiento_produccion(cursor, folio: int) -> Optional[dict]:
    """Lee un movimiento de produccion."""
    cursor.execute("""
        SELECT Folio, Banco, Cuenta, Age, Mes, Dia, Tipo,
               Ingreso, Egreso, Concepto, Clase, FPago, TipoEgreso,
               Conciliada, Paridad, ParidadDOF, Moneda,
               Cia, Fuente, Oficina, CuentaOficina,
               TipoPoliza, NumPoliza, Sucursal
        FROM DBSAV71.dbo.SAVCheqPM
        WHERE Folio = ?
    """, (folio,))
    row = cursor.fetchone()
    if not row:
        return None
    cols = [desc[0] for desc in cursor.description]
    return dict(zip(cols, row))


def _leer_poliza_produccion(cursor, folio: int) -> List[dict]:
    """Lee lineas de poliza de produccion."""
    cursor.execute("""
        SELECT Poliza, Movimiento, Cuenta, SubCuenta,
               TipoCA, Cargo, Abono, Concepto,
               DocTipo, TipoPoliza, DocFolio
        FROM DBSAV71.dbo.SAVPoliza
        WHERE Fuente = 'SAV7-CHEQUES' AND DocFolio = ?
        ORDER BY Poliza, Movimiento
    """, (folio,))
    cols = [desc[0] for desc in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


def _comparar_movimiento(sandbox: dict, produccion: dict) -> Tuple[bool, str]:
    """Compara campos de formato + monto entre sandbox y produccion."""
    diffs = []
    campos = [
        'Banco', 'Cuenta', 'Tipo', 'Moneda', 'Cia', 'Fuente',
        'Oficina', 'CuentaOficina', 'TipoPoliza', 'Sucursal', 'TipoEgreso',
    ]
    for campo in campos:
        val_sb = sandbox.get(campo)
        val_pr = produccion.get(campo)
        if val_sb != val_pr:
            diffs.append(f"{campo}: {val_sb} vs {val_pr}")

    clase_sb = (sandbox.get('Clase') or '').strip()
    clase_pr = (produccion.get('Clase') or '').strip()
    if clase_sb != clase_pr:
        diffs.append(f"Clase: {clase_sb} vs {clase_pr}")

    egreso_sb = Decimal(str(sandbox.get('Egreso', 0)))
    egreso_pr = Decimal(str(produccion.get('Egreso', 0)))
    if abs(egreso_sb - egreso_pr) > Decimal('0.01'):
        diffs.append(f"Egreso: {egreso_sb} vs {egreso_pr}")

    match = len(diffs) == 0
    return match, '; '.join(diffs) if diffs else ''


def _comparar_polizas(
    lineas_sb: List[dict], lineas_pr: List[dict],
) -> Tuple[bool, str]:
    """Compara polizas linea por linea."""
    if len(lineas_sb) != len(lineas_pr):
        return False, f"# lineas: {len(lineas_sb)} vs {len(lineas_pr)}"

    diffs = []
    for i, (ls, lp) in enumerate(zip(lineas_sb, lineas_pr)):
        if ls['Cuenta'] != lp['Cuenta']:
            diffs.append(f"L{i+1} Cuenta: {ls['Cuenta']} vs {lp['Cuenta']}")
        if ls['SubCuenta'] != lp['SubCuenta']:
            diffs.append(f"L{i+1} SubCta: {ls['SubCuenta']} vs {lp['SubCuenta']}")
        if ls['TipoCA'] != lp['TipoCA']:
            diffs.append(f"L{i+1} TipoCA: {ls['TipoCA']} vs {lp['TipoCA']}")

        cargo_sb = Decimal(str(ls['Cargo']))
        cargo_pr = Decimal(str(lp['Cargo']))
        if abs(cargo_sb - cargo_pr) > Decimal('0.01'):
            diffs.append(f"L{i+1} Cargo: {cargo_sb} vs {cargo_pr}")

        abono_sb = Decimal(str(ls['Abono']))
        abono_pr = Decimal(str(lp['Abono']))
        if abs(abono_sb - abono_pr) > Decimal('0.01'):
            diffs.append(f"L{i+1} Abono: {abono_sb} vs {abono_pr}")

    match = len(diffs) == 0
    return match, '; '.join(diffs) if diffs else ''


# ---------------------------------------------------------------------------
# Generador principal — centrado en estado de cuenta
# ---------------------------------------------------------------------------

def generar_reporte_estado_cuenta(
    connector,
    resultados_lineas: List[ResultadoLinea],
    ruta_salida: Path,
):
    """Genera Excel centrado en el estado de cuenta.

    4 hojas: Estado de Cuenta, Resumen por Dia,
    Comparacion Produccion, Polizas.
    """
    logger.info(
        "Generando reporte con {} lineas del estado de cuenta...",
        len(resultados_lineas),
    )

    wb = Workbook()

    # Hoja 1: Estado de Cuenta (eje central)
    ws_ec = wb.active
    ws_ec.title = 'Estado de Cuenta'
    _crear_hoja_estado_cuenta(ws_ec, resultados_lineas)

    # Hoja 2: Resumen por Dia
    ws_dia = wb.create_sheet('Resumen por Dia')
    _crear_hoja_resumen_dia(ws_dia, resultados_lineas)

    # Hojas 3 y 4: Comparacion y Polizas (solo si hay folios y conexion)
    todos_folios = _extraer_folios_unicos(resultados_lineas)

    if todos_folios and connector:
        cursor = connector.db.conectar().cursor()

        ws_comp = wb.create_sheet('Comparacion Produccion')
        _crear_hoja_comparacion(ws_comp, cursor, todos_folios)

        ws_pol = wb.create_sheet('Polizas')
        _crear_hoja_polizas(ws_pol, cursor, todos_folios)

    # Guardar
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    logger.info("Reporte guardado en: {}", ruta_salida)


def _extraer_folios_unicos(
    resultados_lineas: List[ResultadoLinea],
) -> List[Tuple[str, int, bool]]:
    """Extrae (tipo, folio, es_conciliacion) unicos de las lineas."""
    vistos = set()
    items = []
    for rl in resultados_lineas:
        es_conc = rl.accion == AccionLinea.CONCILIAR
        tipo = rl.tipo_clasificado.value
        for folio in rl.folios:
            if folio not in vistos:
                vistos.add(folio)
                items.append((tipo, folio, es_conc))
    return items


# ---------------------------------------------------------------------------
# Hoja 1: Estado de Cuenta
# ---------------------------------------------------------------------------

def _crear_hoja_estado_cuenta(ws, resultados: List[ResultadoLinea]):
    """Cada fila = una linea del estado de cuenta bancario."""
    headers = [
        'Fecha', 'Cuenta', 'Hoja', 'Descripcion',
        'Cargo (Egreso)', 'Abono (Ingreso)',
        'Clasificacion', 'Accion', 'Folio(s)', 'Notas',
    ]

    # Encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN

    # Datos ordenados por fecha, luego por orden original
    for row_idx, rl in enumerate(resultados, 2):
        mov = rl.movimiento
        folios_str = ', '.join(str(f) for f in rl.folios) if rl.folios else ''

        vals = [
            mov.fecha.strftime('%d/%m/%Y'),
            mov.cuenta_banco,
            mov.nombre_hoja,
            mov.descripcion,
            float(mov.cargo) if mov.cargo else None,
            float(mov.abono) if mov.abono else None,
            rl.tipo_clasificado.value,
            rl.accion.value,
            folios_str,
            rl.nota or '',
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN

        # Formato montos
        ws.cell(row=row_idx, column=5).number_format = _FMT_MONEY
        ws.cell(row=row_idx, column=6).number_format = _FMT_MONEY

        # Color de la fila entera segun accion
        fill = _ACCION_FILL.get(rl.accion)
        if fill:
            # Aplicar color a las columnas Accion y Folio(s)
            ws.cell(row=row_idx, column=8).fill = fill
            ws.cell(row=row_idx, column=8).alignment = _ALIGN_CENTER
            if rl.folios:
                ws.cell(row=row_idx, column=9).fill = fill

    ws.freeze_panes = 'A2'
    _autoajustar_columnas(ws)


# ---------------------------------------------------------------------------
# Hoja 2: Resumen por Dia
# ---------------------------------------------------------------------------

def _crear_hoja_resumen_dia(ws, resultados: List[ResultadoLinea]):
    """Una fila por dia con conteos y montos."""
    headers = [
        'Fecha', 'Total Movimientos', 'Procesados (INSERT)',
        'Conciliados', 'Omitidos', 'Sin Procesar',
        'Desconocidos', 'Errores', 'Total Egresos', 'Total Ingresos',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN

    # Agrupar por fecha
    por_dia = defaultdict(list)
    for rl in resultados:
        por_dia[rl.movimiento.fecha].append(rl)

    row_idx = 2
    for fecha in sorted(por_dia.keys()):
        lineas_dia = por_dia[fecha]
        total = len(lineas_dia)
        inserts = sum(1 for rl in lineas_dia if rl.accion == AccionLinea.INSERT)
        conciliados = sum(1 for rl in lineas_dia if rl.accion == AccionLinea.CONCILIAR)
        omitidos = sum(1 for rl in lineas_dia if rl.accion == AccionLinea.OMITIR)
        sin_proc = sum(1 for rl in lineas_dia if rl.accion == AccionLinea.SIN_PROCESAR)
        desconocidos = sum(1 for rl in lineas_dia if rl.accion == AccionLinea.DESCONOCIDO)
        errores = sum(1 for rl in lineas_dia if rl.accion == AccionLinea.ERROR)

        total_egresos = sum(
            float(rl.movimiento.cargo)
            for rl in lineas_dia
            if rl.movimiento.cargo
        )
        total_ingresos = sum(
            float(rl.movimiento.abono)
            for rl in lineas_dia
            if rl.movimiento.abono
        )

        vals = [
            fecha.strftime('%d/%m/%Y'),
            total, inserts, conciliados, omitidos,
            sin_proc, desconocidos, errores,
            total_egresos, total_ingresos,
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN
            if row_idx % 2 == 0:
                cell.fill = _FILL_STRIPE

        ws.cell(row=row_idx, column=9).number_format = _FMT_MONEY
        ws.cell(row=row_idx, column=10).number_format = _FMT_MONEY

        # Destacar desconocidos > 0
        if desconocidos > 0:
            ws.cell(row=row_idx, column=7).fill = _FILL_DESCONOCIDO

        row_idx += 1

    # Fila de totales
    if por_dia:
        row_total = row_idx
        ws.cell(row=row_total, column=1, value='TOTAL').font = Font(bold=True, size=10)
        for col in range(2, 11):
            total = sum(
                ws.cell(row=r, column=col).value or 0
                for r in range(2, row_idx)
            )
            cell = ws.cell(row=row_total, column=col, value=total)
            cell.font = Font(bold=True, size=10)
            cell.border = _BORDER_THIN
        ws.cell(row=row_total, column=9).number_format = _FMT_MONEY
        ws.cell(row=row_total, column=10).number_format = _FMT_MONEY

    ws.freeze_panes = 'A2'
    _autoajustar_columnas(ws)


# ---------------------------------------------------------------------------
# Hoja 3: Comparacion Produccion
# ---------------------------------------------------------------------------

def _crear_hoja_comparacion(ws, cursor, items: List[Tuple[str, int, bool]]):
    """Folios creados en sandbox vs produccion."""
    headers = [
        'Tipo Proceso', 'Fecha', 'Folio Sandbox', 'Clase', 'Concepto',
        'Egreso', 'Ingreso', 'TipoPoliza', '# Lineas Poliza',
        'Folio Produccion', 'Match Movimiento', 'Match Poliza', 'Diferencias',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN

    for row_idx, (tipo, folio, es_conc) in enumerate(items, 2):
        mov = _leer_movimiento(cursor, folio)
        if not mov:
            ws.cell(row=row_idx, column=1, value=tipo)
            ws.cell(row=row_idx, column=3, value=folio)
            ws.cell(row=row_idx, column=13, value='Movimiento no encontrado')
            continue

        poliza_sb = _leer_poliza(cursor, folio)
        fecha = date(mov['Age'], mov['Mes'], mov['Dia'])
        clase = (mov.get('Clase') or '').strip()
        egreso = float(mov.get('Egreso', 0))

        folio_prod = None
        match_mov = 'N/A'
        match_pol = 'N/A'
        diffs_texto = ''

        if not es_conc:
            folio_prod = _buscar_produccion(
                cursor, clase, egreso, mov['Age'], mov['Mes'],
            )
            if folio_prod:
                mov_prod = _leer_movimiento_produccion(cursor, folio_prod)
                poliza_prod = _leer_poliza_produccion(cursor, folio_prod)

                if mov_prod:
                    ok_mov, diffs_mov = _comparar_movimiento(mov, mov_prod)
                    match_mov = 'SI' if ok_mov else 'NO'
                    if diffs_mov:
                        diffs_texto = diffs_mov

                if poliza_sb and poliza_prod:
                    ok_pol, diffs_pol = _comparar_polizas(poliza_sb, poliza_prod)
                    match_pol = 'SI' if ok_pol else 'NO'
                    if diffs_pol:
                        diffs_texto = (diffs_texto + ' | ' + diffs_pol).strip(' | ')
            else:
                diffs_texto = 'Sin match en produccion'

        vals = [
            tipo,
            fecha.strftime('%d/%m/%Y'),
            folio,
            clase,
            (mov.get('Concepto') or '').strip(),
            egreso,
            float(mov.get('Ingreso', 0)),
            (mov.get('TipoPoliza') or '').strip(),
            len(poliza_sb),
            folio_prod or '',
            match_mov,
            match_pol,
            diffs_texto,
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = _FONT_NORMAL
            cell.border = _BORDER_THIN
            if row_idx % 2 == 0:
                cell.fill = _FILL_STRIPE

        ws.cell(row=row_idx, column=6).number_format = _FMT_MONEY
        ws.cell(row=row_idx, column=7).number_format = _FMT_MONEY

        _aplicar_color_match(ws.cell(row=row_idx, column=11), match_mov)
        _aplicar_color_match(ws.cell(row=row_idx, column=12), match_pol)

    ws.freeze_panes = 'A2'
    _autoajustar_columnas(ws)


# ---------------------------------------------------------------------------
# Hoja 4: Polizas
# ---------------------------------------------------------------------------

def _crear_hoja_polizas(ws, cursor, items: List[Tuple[str, int, bool]]):
    """Comparacion linea por linea de polizas sandbox vs produccion."""
    headers = [
        'Folio Sandbox', 'Tipo Proceso', 'Linea', 'Cuenta', 'SubCuenta',
        'TipoCA', 'Cargo Sandbox', 'Abono Sandbox',
        'Cargo Produccion', 'Abono Produccion', 'Match',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_THIN

    row_idx = 2

    for tipo, folio, es_conc in items:
        if es_conc:
            continue

        poliza_sb = _leer_poliza(cursor, folio)
        if not poliza_sb:
            continue

        mov = _leer_movimiento(cursor, folio)
        poliza_prod = []
        if mov:
            clase = (mov.get('Clase') or '').strip()
            egreso = float(mov.get('Egreso', 0))
            folio_prod = _buscar_produccion(
                cursor, clase, egreso, mov['Age'], mov['Mes'],
            )
            if folio_prod:
                poliza_prod = _leer_poliza_produccion(cursor, folio_prod)

        for i, ls in enumerate(poliza_sb):
            lp = poliza_prod[i] if i < len(poliza_prod) else None

            cargo_sb = float(ls['Cargo'])
            abono_sb = float(ls['Abono'])
            cargo_pr = float(lp['Cargo']) if lp else None
            abono_pr = float(lp['Abono']) if lp else None

            match = 'N/A'
            if lp:
                cuenta_ok = ls['Cuenta'] == lp['Cuenta']
                subcta_ok = ls['SubCuenta'] == lp['SubCuenta']
                tipoca_ok = ls['TipoCA'] == lp['TipoCA']
                cargo_ok = abs(
                    Decimal(str(ls['Cargo'])) - Decimal(str(lp['Cargo']))
                ) <= Decimal('0.01')
                abono_ok = abs(
                    Decimal(str(ls['Abono'])) - Decimal(str(lp['Abono']))
                ) <= Decimal('0.01')
                match = 'SI' if all([
                    cuenta_ok, subcta_ok, tipoca_ok, cargo_ok, abono_ok,
                ]) else 'NO'

            tipoca_str = 'CARGO' if ls['TipoCA'] == 1 else 'ABONO'

            vals = [
                folio, tipo, i + 1,
                ls['Cuenta'], ls['SubCuenta'], tipoca_str,
                cargo_sb, abono_sb,
                cargo_pr, abono_pr,
                match,
            ]

            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = _FONT_NORMAL
                cell.border = _BORDER_THIN

            for c in (7, 8, 9, 10):
                ws.cell(row=row_idx, column=c).number_format = _FMT_MONEY

            _aplicar_color_match(ws.cell(row=row_idx, column=11), match)

            row_idx += 1

    ws.freeze_panes = 'A2'
    _autoajustar_columnas(ws)


# ---------------------------------------------------------------------------
# Funcion legacy (compatibilidad con demo anterior)
# ---------------------------------------------------------------------------

def generar_reporte_demo(
    connector,
    resultados: List[ResultadoProceso],
    ruta_salida: Path,
):
    """Genera el Excel de demo con comparacion vs produccion (version legacy)."""
    cursor = connector.db.conectar().cursor()
    items = _extraer_folios_legacy(resultados)

    if not items:
        logger.warning("Sin folios para reportar")
        return

    wb = Workbook()
    ws_res = wb.active
    ws_res.title = 'Resumen'
    _crear_hoja_comparacion(ws_res, cursor, items)

    ws_pol = wb.create_sheet('Polizas')
    _crear_hoja_polizas(ws_pol, cursor, items)

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    logger.info("Reporte guardado en: {}", ruta_salida)


def _extraer_folios_legacy(
    resultados: List[ResultadoProceso],
) -> List[Tuple[str, int, bool]]:
    """Extrae folios de ResultadoProceso (version legacy)."""
    items = []
    for r in resultados:
        if not r.exito:
            continue
        for folio in r.folios:
            items.append((r.tipo_proceso, folio, False))
        if r.plan and r.plan.conciliaciones:
            for conc in r.plan.conciliaciones:
                folio = conc.get('folio')
                if folio:
                    items.append((r.tipo_proceso, folio, True))
    return items


# ---------------------------------------------------------------------------
# Utilidades de formato
# ---------------------------------------------------------------------------

def _aplicar_color_match(cell, valor: str):
    """Aplica color de fondo segun valor de match."""
    if valor == 'SI':
        cell.fill = _FILL_SI
    elif valor == 'NO':
        cell.fill = _FILL_NO
    elif valor == 'N/A':
        cell.fill = _FILL_NA
    cell.alignment = _ALIGN_CENTER


def _autoajustar_columnas(ws):
    """Ajusta el ancho de columnas al contenido (aproximado)."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        adjusted = min(max_length + 3, 50)
        ws.column_dimensions[col_letter].width = max(adjusted, 10)
