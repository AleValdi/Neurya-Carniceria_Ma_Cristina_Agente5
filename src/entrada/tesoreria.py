"""Parser del reporte de tesoreria (ingresos diarios).

Lee archivos como 'FEBRERO INGRESOS 2026.xlsx' que contienen una hoja
por dia del mes. Estructura esperada por hoja:
  - J18: Fecha del corte (autoritativa, NO usar E3). Fallback: numero de hoja = dia
  - G19:G43 + H19:H43: Facturas individuales (numero + importe)
  - K19 o K20: Numero de factura global (posicion variable!)
  - L20: Importe de factura global
  - D44: Total ventas (suma cortes Z)
  - E63: Total efectivo recibido
  - H63: Total TDC (4 terminales)
  - L55: Total otros medios de pago
  - D65: Folio SISSA
"""

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from loguru import logger

from src.entrada.normalizacion import parsear_fecha_excel, normalizar_monto
from src.models import CorteVentaDiaria, FacturaVenta


def parsear_tesoreria(ruta: Path) -> Dict[date, CorteVentaDiaria]:
    """Parsea el reporte de tesoreria completo.

    Args:
        ruta: Ruta al archivo Excel de ingresos.

    Returns:
        Dict con clave = fecha del corte, valor = CorteVentaDiaria.
        Solo incluye hojas que tienen datos. Usa J18 como fecha; si falta,
        usa el numero de hoja como dia del mes (Hoja 1 = Dia 1).
    """
    logger.info("Parseando reporte de tesoreria: {}", ruta.name)
    wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=False)

    resultado: Dict[date, CorteVentaDiaria] = {}
    hojas_sin_fecha: List[tuple] = []  # (indice, nombre_hoja, worksheet)

    for idx, nombre_hoja in enumerate(wb.sheetnames):
        ws = wb[nombre_hoja]
        corte = _parsear_hoja_diaria(ws, nombre_hoja)

        if corte is not None and _tiene_datos(corte):
            resultado[corte.fecha_corte] = corte
            _log_corte(nombre_hoja, corte)
        elif corte is None:
            # J18 sin fecha — guardar para fallback por numero de hoja
            hojas_sin_fecha.append((idx, nombre_hoja, ws))

    # Fallback: numero de hoja = dia del mes (Hoja 0 = Dia 1)
    if hojas_sin_fecha and resultado:
        fecha_ref = next(iter(resultado.values())).fecha_corte
        for idx, nombre_hoja, ws in hojas_sin_fecha:
            dia = idx + 1
            try:
                fecha_fallback = date(fecha_ref.year, fecha_ref.month, dia)
            except ValueError:
                continue  # dia invalido para el mes
            corte = _parsear_hoja_diaria(ws, nombre_hoja, fecha_fallback)
            if corte is not None and _tiene_datos(corte):
                resultado[corte.fecha_corte] = corte
                _log_corte(nombre_hoja, corte)

    wb.close()

    logger.info(
        "Tesoreria: {} dias con datos de {} hojas totales",
        len(resultado), len(wb.sheetnames),
    )
    return resultado


def _log_corte(nombre_hoja: str, corte: CorteVentaDiaria):
    """Logea informacion de un corte parseado."""
    n_ind = len(corte.facturas_individuales)
    logger.info(
        "Hoja '{}': corte {} | {} fact.indiv | global={} ${} | "
        "efectivo=${} | TDC=${}",
        nombre_hoja,
        corte.fecha_corte,
        n_ind,
        corte.factura_global_numero or '-',
        corte.factura_global_importe or 0,
        corte.total_efectivo or 0,
        corte.total_tdc or 0,
    )


def _parsear_hoja_diaria(
    ws, nombre_hoja: str, fecha_fallback: Optional[date] = None,
) -> Optional[CorteVentaDiaria]:
    """Parsea una hoja diaria del reporte de tesoreria.

    Args:
        ws: Worksheet de openpyxl.
        nombre_hoja: Nombre de la hoja (para logs).
        fecha_fallback: Fecha a usar si J18 no tiene fecha valida.
            Se calcula como numero_de_hoja = dia_del_mes.
    """
    # Fecha del corte: J18 (autoritativa), fallback por numero de hoja
    fecha_corte = parsear_fecha_excel(_celda(ws, 'J', 18))
    if fecha_corte is None:
        if fecha_fallback is not None:
            fecha_corte = fecha_fallback
            logger.warning(
                "Hoja '{}': sin fecha en J18, usando dia {} por numero de hoja",
                nombre_hoja, fecha_fallback.day,
            )
        else:
            logger.debug("Hoja '{}': sin fecha en J18, saltando", nombre_hoja)
            return None

    corte = CorteVentaDiaria(
        fecha_corte=fecha_corte,
        nombre_hoja=nombre_hoja,
    )

    # Facturas individuales: G19:G43 (numero) + H19:H43 (importe)
    corte.facturas_individuales = _parsear_facturas_individuales(ws)

    # Factura global: K19 o K20 (posicion variable!) + L20 (importe)
    corte.factura_global_numero = _parsear_factura_global_numero(ws)
    corte.factura_global_importe = normalizar_monto(_celda(ws, 'L', 20))

    # Totales
    corte.total_ventas = normalizar_monto(_celda(ws, 'D', 44))
    corte.total_efectivo = normalizar_monto(_celda(ws, 'E', 63))
    corte.total_tdc = normalizar_monto(_celda(ws, 'H', 63))
    corte.total_otros = normalizar_monto(_celda(ws, 'L', 55))
    corte.folio_sissa = _celda_str(ws, 'D', 65)

    return corte


def _parsear_facturas_individuales(ws) -> List[FacturaVenta]:
    """Extrae facturas individuales de G19:G43 + H19:H43."""
    facturas = []

    for fila in range(19, 44):  # 19 a 43 inclusive
        numero_raw = _celda(ws, 'G', fila)
        importe_raw = _celda(ws, 'H', fila)

        if numero_raw is None or importe_raw is None:
            continue

        importe = normalizar_monto(importe_raw)
        if importe is None or importe == Decimal('0'):
            continue

        # El numero puede ser int o string
        numero = str(int(numero_raw)) if isinstance(numero_raw, (int, float)) else str(numero_raw).strip()
        if not numero:
            continue

        facturas.append(FacturaVenta(
            serie='FD',
            numero=numero,
            importe=importe,
        ))

    return facturas


def _parsear_factura_global_numero(ws) -> Optional[str]:
    """Extrae el numero de factura global de K19 o K20.

    Anomalia documentada: en hojas 1-13 esta en K19, en hojas 14-15 en K20.
    Checamos ambas posiciones.
    """
    for fila in (19, 20):
        valor = _celda(ws, 'K', fila)
        if valor is not None:
            numero = str(int(valor)) if isinstance(valor, (int, float)) else str(valor).strip()
            if numero:
                return numero
    return None


def _tiene_datos(corte: CorteVentaDiaria) -> bool:
    """Verifica si un corte tiene datos reales (no solo fecha)."""
    if corte.facturas_individuales:
        return True
    if corte.factura_global_numero:
        return True
    if corte.total_efectivo and corte.total_efectivo > 0:
        return True
    if corte.total_tdc and corte.total_tdc > 0:
        return True
    return False


def _celda(ws, col: str, fila: int):
    """Obtiene el valor de una celda por referencia (ej: 'J', 18)."""
    return ws[f'{col}{fila}'].value


def _celda_str(ws, col: str, fila: int) -> Optional[str]:
    """Obtiene el valor de una celda como string."""
    valor = _celda(ws, col, fila)
    if valor is None:
        return None
    return str(valor).strip() or None
