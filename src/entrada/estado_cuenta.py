"""Parser del estado de cuenta bancario (Excel Banregio).

Lee archivos como PRUEBA.xlsx que contienen hojas por cuenta bancaria.
Estructura esperada por hoja:
  - Filas 1-4: Encabezado (razon social, cuenta, CLABE, RFC, saldos)
  - Fila 5: Headers (Fecha | Descripcion/Referencia | Cargos | Abonos | Saldo)
  - Fila 6+: Datos de movimientos
  - Fin de datos: donde columna A deja de tener datetime
"""

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from loguru import logger

from config.settings import HOJAS_ESTADO_CUENTA, CUENTAS_BANCARIAS
from src.entrada.normalizacion import fix_mojibake, parsear_fecha_excel, normalizar_monto
from src.models import MovimientoBancario


# Fila donde empiezan los datos (1-indexed)
FILA_DATOS_INICIO = 6

# Columnas (1-indexed): A=Fecha, B=Descripcion, C=Cargos, D=Abonos, E=Saldo
COL_FECHA = 1
COL_DESCRIPCION = 2
COL_CARGOS = 3
COL_ABONOS = 4


def parsear_estado_cuenta(
    ruta: Path,
) -> Dict[str, List[MovimientoBancario]]:
    """Parsea un archivo de estado de cuenta bancario.

    Args:
        ruta: Ruta al archivo Excel.

    Returns:
        Dict con clave = nombre de hoja, valor = lista de MovimientoBancario.
        Solo incluye hojas reconocidas (definidas en HOJAS_ESTADO_CUENTA).
    """
    logger.info("Parseando estado de cuenta: {}", ruta.name)
    wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)

    resultado: Dict[str, List[MovimientoBancario]] = {}

    for nombre_hoja in wb.sheetnames:
        clave_cuenta = _identificar_hoja(nombre_hoja)
        if clave_cuenta is None:
            logger.debug("Hoja '{}' no reconocida, saltando", nombre_hoja)
            continue

        cuenta_config = CUENTAS_BANCARIAS[clave_cuenta]
        ws = wb[nombre_hoja]

        movimientos = _parsear_hoja(ws, nombre_hoja, cuenta_config.cuenta)
        if movimientos:
            resultado[nombre_hoja] = movimientos
            logger.info(
                "Hoja '{}': {} movimientos parseados (cuenta {})",
                nombre_hoja, len(movimientos), cuenta_config.cuenta,
            )

    wb.close()

    total = sum(len(m) for m in resultado.values())
    logger.info("Total: {} movimientos en {} hojas", total, len(resultado))

    return resultado


def parsear_estado_cuenta_plano(ruta: Path) -> List[MovimientoBancario]:
    """Parsea y retorna todos los movimientos en una sola lista."""
    por_hoja = parsear_estado_cuenta(ruta)
    todos = []
    for movimientos in por_hoja.values():
        todos.extend(movimientos)
    return todos


def _identificar_hoja(nombre_hoja: str) -> Optional[str]:
    """Identifica la clave de cuenta para una hoja del estado de cuenta."""
    # Buscar match exacto primero
    if nombre_hoja in HOJAS_ESTADO_CUENTA:
        return HOJAS_ESTADO_CUENTA[nombre_hoja]

    # Buscar match por strip
    nombre_limpio = nombre_hoja.strip()
    for nombre_config, clave in HOJAS_ESTADO_CUENTA.items():
        if nombre_limpio == nombre_config.strip():
            return clave

    return None


def _parsear_hoja(
    ws,
    nombre_hoja: str,
    cuenta_banco: str,
) -> List[MovimientoBancario]:
    """Parsea una hoja individual del estado de cuenta."""
    movimientos: List[MovimientoBancario] = []

    for fila in ws.iter_rows(min_row=FILA_DATOS_INICIO, values_only=False):
        # Columna A (fecha): si no es datetime/date, fin de datos
        celda_fecha = fila[COL_FECHA - 1]
        fecha = parsear_fecha_excel(celda_fecha.value)
        if fecha is None:
            break

        # Columna B (descripcion)
        descripcion_raw = fila[COL_DESCRIPCION - 1].value
        if descripcion_raw is None:
            descripcion_raw = ''
        descripcion = fix_mojibake(str(descripcion_raw))

        # Columna C (cargos/egresos)
        cargo = normalizar_monto(fila[COL_CARGOS - 1].value)

        # Columna D (abonos/ingresos)
        abono = normalizar_monto(fila[COL_ABONOS - 1].value)

        # Saltar filas sin monto
        if cargo is None and abono is None:
            continue

        movimiento = MovimientoBancario(
            fecha=fecha,
            descripcion=descripcion,
            cargo=cargo,
            abono=abono,
            cuenta_banco=cuenta_banco,
            nombre_hoja=nombre_hoja,
        )
        movimientos.append(movimiento)

    return movimientos


def obtener_metadata_hoja(
    ruta: Path,
    nombre_hoja: str,
) -> Dict[str, str]:
    """Extrae metadata del encabezado de una hoja (filas 1-4).

    Retorna dict con claves como 'razon_social', 'cuenta', 'clabe', etc.
    """
    wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
    ws = wb[nombre_hoja]

    metadata = {}
    filas = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))

    if len(filas) >= 1 and filas[0][0]:
        metadata['razon_social'] = str(filas[0][0]).strip()
    if len(filas) >= 2 and filas[1][0]:
        metadata['cuenta'] = str(filas[1][0]).strip()
    if len(filas) >= 3 and filas[2][0]:
        metadata['clabe'] = str(filas[2][0]).strip()
    if len(filas) >= 4 and filas[3][0]:
        metadata['rfc'] = str(filas[3][0]).strip()

    wb.close()
    return metadata
