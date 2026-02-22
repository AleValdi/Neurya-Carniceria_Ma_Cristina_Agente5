"""Parser para ajustes opcionales de impuestos desde Excel.

Si el archivo AJUSTES_IMPUESTOS.xlsx existe en data/reportes/,
los valores que contenga sobreescriben los extraidos de los PDFs.
Celdas vacias se ignoran (se usa el valor del PDF).

Estructura esperada:
    A1: Concepto          B1: Importe
    A2: IMSS              B2: (monto o vacio)
    A3: IVA Acumulable    B3: (monto o vacio)
    A4: IVA Acreditable   B4: (monto o vacio)
    A5: (vacio)
    A6: Proveedor         B6: Nombre           C6: Monto
    A7: 001640            B7: AUTOTRANSPORTE   C7: 154.00
    A8: ...               B8: ...              C8: ...
"""

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, List, Optional

from loguru import logger
import openpyxl

from src.models import RetencionIVAProveedor


def parsear_ajustes_impuestos(ruta: Path) -> Dict:
    """Lee ajustes opcionales de impuestos desde Excel.

    Returns:
        Dict con claves presentes solo si tienen valor:
        - 'total_imss': Decimal
        - 'iva_acumulable': Decimal
        - 'iva_acreditable': Decimal
        - 'retenciones_iva': List[RetencionIVAProveedor]  (si hay filas 7+)
    """
    MAPA_MONTOS = {
        2: 'total_imss',
        3: 'iva_acumulable',
        4: 'iva_acreditable',
    }
    FILA_INICIO_RETENCIONES = 7  # Primera fila de datos de retenciones

    try:
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb.active

        ajustes = {}

        # --- Montos fijos (B2, B3, B4) ---
        for fila, clave in MAPA_MONTOS.items():
            valor = ws.cell(row=fila, column=2).value
            if valor is not None:
                try:
                    ajustes[clave] = Decimal(str(valor))
                except (InvalidOperation, ValueError):
                    logger.warning(
                        "Ajuste impuestos: celda B{} no es numerico ({}), ignorando",
                        fila, valor,
                    )

        # --- Retenciones IVA por proveedor (fila 7+) ---
        retenciones = []
        for fila in range(FILA_INICIO_RETENCIONES, ws.max_row + 1):
            proveedor = ws.cell(row=fila, column=1).value
            nombre = ws.cell(row=fila, column=2).value
            monto_raw = ws.cell(row=fila, column=3).value

            if not proveedor or monto_raw is None:
                continue

            try:
                monto = Decimal(str(monto_raw))
            except (InvalidOperation, ValueError):
                logger.warning(
                    "Ajuste retenciones: fila {} monto no numerico ({}), ignorando",
                    fila, monto_raw,
                )
                continue

            retenciones.append(RetencionIVAProveedor(
                proveedor=str(proveedor).strip(),
                nombre=str(nombre or '').strip(),
                monto=monto,
            ))

        if retenciones:
            ajustes['retenciones_iva'] = retenciones

        wb.close()
        return ajustes

    except Exception as e:
        logger.warning("Error leyendo ajustes de impuestos {}: {}", ruta, e)
        return {}
