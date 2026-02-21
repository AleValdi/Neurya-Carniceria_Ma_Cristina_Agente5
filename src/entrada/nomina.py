"""Parser del archivo de nomina CONTPAQi.

Lee archivos como 'NOMINA 03 CHEQUE.xlsx' con 4 hojas:
  - NOM 03: Nomina completa (empleados + totales)
  - DISPERSION: Transferencias bancarias
  - CHEQUE: Pagos en cheque
  - CHEQUE FINIQUITOS: Finiquitos por cheque

Estructura de hoja NOM XX:
  - Filas 73-78: Totales (DISPERSION, CHEQUES, etc.)
  - Fila 81: Vacaciones pagadas
  - Fila 84: Finiquito

Tambien soporta parsear percepciones/deducciones desde el PDF
'Lista de raya XX.pdf' generado por CONTPAQi Nominas.
"""

import re
import unicodedata
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from loguru import logger

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

from src.entrada.normalizacion import normalizar_monto
from src.models import DatosNomina, LineaContable, MovimientoNomina


def _normalizar_texto(texto: str) -> str:
    """Normaliza texto: quita acentos, puntos, y pasa a mayusculas.

    Ej: 'Séptimo día' → 'SEPTIMO DIA'
        'I.S.R. (mes)' → 'ISR (MES)'
        'Préstamo infonavit (FD)' → 'PRESTAMO INFONAVIT (FD)'
    """
    # Quitar acentos
    nfkd = unicodedata.normalize('NFKD', texto)
    sin_acentos = ''.join(c for c in nfkd if not unicodedata.combining(c))
    # Quitar puntos
    sin_puntos = sin_acentos.replace('.', '')
    return sin_puntos.upper().strip()


# Mapeo de percepciones: concepto → (cuenta, subcuenta)
PERCEPCIONES_CUENTAS = {
    'SUELDO': ('6200', '010000'),
    'SUELDOS': ('6200', '010000'),
    'SEPTIMO DIA': ('6200', '240000'),
    'PRIMA DOMINICAL': ('6200', '670000'),
    'BONO PUNTUALIDAD': ('6200', '770000'),
    'BONO DE PUNTUALIDAD': ('6200', '770000'),
    'VACACIONES': ('6200', '020000'),
    'PRIMA VACACIONAL': ('6200', '060000'),
    'AGUINALDO': ('6200', '030000'),
    'BONO ASISTENCIA': ('6200', '780000'),
    'BONO DE ASISTENCIA': ('6200', '780000'),
}

# Mapeo de deducciones: concepto → (cuenta, subcuenta)
DEDUCCIONES_CUENTAS = {
    'INFONAVIT VIVIENDA': ('2140', '270000'),
    'INFONAVIT FD': ('2140', '270000'),
    'INFONAVIT (FD)': ('2140', '270000'),
    'INFONAVIT CF': ('2140', '270000'),
    'INFONAVIT (CF)': ('2140', '270000'),
    'ISR': ('2140', '020000'),
    'ISR (MES)': ('2140', '020000'),
    'IMSS': ('2140', '010000'),
    'I.M.S.S.': ('2140', '010000'),
}


# --- Mapeos para PDF Lista de Raya ---

# Percepciones conocidas del PDF (concepto normalizado → cuenta contable)
PERCEPCIONES_PDF = {
    'SUELDO': ('6200', '010000'),
    'SUELDOS': ('6200', '010000'),
    'SEPTIMO DIA': ('6200', '240000'),
    'PRIMA DOMINICAL': ('6200', '670000'),
    'BONO PUNTUALIDAD': ('6200', '770000'),
    'BONO DE PUNTUALIDAD': ('6200', '770000'),
    'VACACIONES A TIEMPO': ('6200', '020000'),
    'PRIMA DE VACACIONES A TIEMPO': ('6200', '060000'),
    'VACACIONES REPORTADAS': ('6200', '020000'),
    'VACACIONES REPORTADAS $': ('6200', '020000'),
    'AGUINALDO': ('6200', '030000'),
    'BONO DE ASISTENCIA': ('6200', '780000'),
    'BONO ASISTENCIA': ('6200', '780000'),
    'GRATIFICACIONES': ('6200', '260000'),
    'INDEMNIZACIONES': ('6200', '270000'),
}

# Deducciones reales del PDF (no informativas)
DEDUCCIONES_PDF = {
    'SEGURO DE VIVIENDA INFONAVIT': ('2140', '270000'),
    'PRESTAMO INFONAVIT (FD)': ('2140', '270000'),
    'PRESTAMO INFONAVIT (CF)': ('2140', '270000'),
    'ISR (MES)': ('2140', '020000'),
    'IMSS': ('2140', '010000'),
}

# Numeros de concepto a omitir: informativos (32,41,99,180,181) + obligaciones (89-98)
_CONCEPTOS_OMITIR = {32, 41, 89, 90, 93, 96, 97, 98, 99, 180, 181}

# Regex para extraer entradas: numero + concepto + monto
_RE_ENTRADA_PDF = re.compile(
    r'(\d{1,3})\s+'
    r'([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ\s\.\(\)\$]+?)\s+'
    r'(-?[\d,]+\.\d{2})'
)


def parsear_nomina(ruta: Path, ruta_lista_raya: Path = None) -> Optional[DatosNomina]:
    """Parsea un archivo de nomina CONTPAQi.

    Args:
        ruta: Ruta al archivo Excel de nomina.
        ruta_lista_raya: Ruta al PDF Lista de Raya (opcional).
            Si se proporciona, las percepciones/deducciones se extraen
            del PDF en vez del Excel.

    Returns:
        DatosNomina con totales y detalle de percepciones/deducciones,
        o None si no se pudo parsear.
    """
    logger.info("Parseando nomina: {}", ruta.name)

    # Extraer numero de nomina del nombre del archivo
    numero_nomina = _extraer_numero_nomina(ruta.name)

    wb = openpyxl.load_workbook(str(ruta), data_only=True)

    # Buscar hoja de nomina principal (NOM XX)
    hoja_nomina = _buscar_hoja_nomina(wb)
    if hoja_nomina is None:
        logger.error("No se encontro hoja de nomina en {}", ruta.name)
        wb.close()
        return None

    ws = wb[hoja_nomina]

    # Extraer movimientos dinamicamente del Excel
    movimientos = _extraer_movimientos(ws)

    # Percepciones y deducciones: preferir PDF si disponible
    if ruta_lista_raya and ruta_lista_raya.exists():
        percepciones, deducciones = parsear_lista_raya_pdf(ruta_lista_raya)
        logger.info("  Usando Lista de Raya PDF para percepciones/deducciones")
    else:
        percepciones, deducciones = _extraer_percepciones_deducciones(ws)

    wb.close()

    datos = DatosNomina(
        numero_nomina=numero_nomina,
        movimientos=movimientos,
        percepciones=percepciones,
        deducciones=deducciones,
    )

    logger.info(
        "Nomina {}: {} movimientos, neto=${:,.2f}",
        numero_nomina, len(movimientos), datos.total_neto,
    )
    for mov in movimientos:
        logger.info(
            "  {} ${:,.2f} (clase={}, egreso={})",
            mov.tipo, mov.monto, mov.clase, mov.tipo_egreso,
        )

    return datos


def _extraer_numero_nomina(nombre_archivo: str) -> int:
    """Extrae el numero de nomina del nombre del archivo.

    Ej: 'NOMINA 03 CHEQUE.xlsx' → 3
    """
    match = re.search(r'NOMINA\s+(\d+)', nombre_archivo, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return 0


def _buscar_hoja_nomina(wb) -> Optional[str]:
    """Busca la hoja principal de nomina (NOM XX)."""
    for nombre in wb.sheetnames:
        if re.match(r'NOM\s+\d+', nombre, re.IGNORECASE):
            return nombre
    # Fallback: primera hoja
    if wb.sheetnames:
        return wb.sheetnames[0]
    return None


def _extraer_movimientos(ws) -> List[MovimientoNomina]:
    """Descubre movimientos de nomina dinamicamente.

    Dos fuentes:
      1. Columna I: DISPERSION y CHEQUES (subtotales del payroll, filas ~74-75)
      2. Columna O: movimientos adicionales (VAC PAGADAS, FINIQUITO PAGADO, etc.)

    Las filas pueden variar entre periodos.  El numero y tipo de movimientos
    tambien puede variar.

    Returns:
        Lista de MovimientoNomina ordenada: principal primero, luego secundarios.
    """
    movimientos = []

    # --- Fase 1: DISPERSION y CHEQUES desde columna I ---
    for fila in range(70, 91):
        etiqueta_i = _celda_str(ws, 'I', fila)
        if not etiqueta_i:
            continue

        etiqueta_upper = etiqueta_i.upper().strip()
        monto = normalizar_monto(_celda(ws, 'J', fila))
        if monto is None or monto <= 0:
            continue

        if 'DISPER' in etiqueta_upper:
            movimientos.append(MovimientoNomina(
                tipo='DISPERSION', monto=monto,
                clase='NOMINA', tipo_egreso='TRANSFERENCIA',
                es_principal=True,
            ))
        elif 'CHEQUE' in etiqueta_upper and 'FINIQ' not in etiqueta_upper:
            movimientos.append(MovimientoNomina(
                tipo='CHEQUES', monto=monto,
                clase='NOMINA', tipo_egreso='CHEQUE',
                es_principal=False,
            ))

    # --- Fase 2: movimientos adicionales desde columna O ---
    max_fila = min(ws.max_row or 150, 200)
    for fila in range(70, max_fila + 1):
        val_o = _celda_str(ws, 'O', fila)
        if not val_o:
            continue

        val_o_upper = val_o.upper().strip()
        # Omitir header y ceros
        if val_o_upper in ('DETALLE', '0', ''):
            continue

        monto = normalizar_monto(_celda(ws, 'J', fila))
        if monto is None or monto <= 0:
            continue

        clase, tipo_egreso = _clasificar_movimiento_nomina(val_o_upper)
        movimientos.append(MovimientoNomina(
            tipo=val_o_upper, monto=monto,
            clase=clase, tipo_egreso=tipo_egreso,
            es_principal=False,
        ))

    # Asegurar que principal va primero
    movimientos.sort(key=lambda m: (not m.es_principal, m.tipo))

    return movimientos


def _clasificar_movimiento_nomina(tipo: str) -> Tuple[str, str]:
    """Determina Clase y TipoEgreso a partir del tipo de movimiento.

    Returns:
        Tupla (clase, tipo_egreso).
    """
    tipo_upper = tipo.upper()
    if 'FINIQ' in tipo_upper:
        return 'FINIQUITO', 'TRANSFERENCIA'
    return 'NOMINA', 'TRANSFERENCIA'


def _extraer_percepciones_deducciones(ws) -> tuple:
    """Extrae percepciones y deducciones de la nomina.

    Estructura del Excel CONTPAQi (NOM XX):
      - Fila 5: Headers de columnas (C=Sueldo, D=Séptimo día, etc.)
      - Fila 73: Totales por columna (suma de todos los empleados)

    Lee los headers de fila 5, busca conceptos conocidos en PERCEPCIONES_CUENTAS
    y DEDUCCIONES_CUENTAS, y extrae los totales de fila 73.

    Returns:
        Tupla (percepciones, deducciones) como listas de LineaContable.
    """
    percepciones = []
    deducciones = []

    # Buscar la fila de totales: primera fila despues de los empleados
    # donde columna A no tiene numero de empleado y columna C tiene valor
    fila_totales = None
    for fila in range(70, 91):
        val_a = ws[f'A{fila}'].value
        val_c = ws[f'C{fila}'].value
        # Fila de totales: no tiene codigo de empleado en A, pero tiene montos en C+
        if val_a is None and val_c is not None:
            monto = normalizar_monto(val_c)
            if monto is not None and monto > 0:
                fila_totales = fila
                break

    if fila_totales is None:
        return percepciones, deducciones

    # Leer headers de fila 5 y totales de la fila encontrada
    for col in 'CDEFGHIJK':
        header = _celda_str(ws, col, 5)
        if not header:
            continue

        monto = normalizar_monto(_celda(ws, col, fila_totales))
        if monto is None or monto <= 0:
            continue

        header_norm = _normalizar_texto(header)

        # Buscar en percepciones
        cuenta_info = None
        for concepto, cta in PERCEPCIONES_CUENTAS.items():
            if concepto in header_norm or header_norm in concepto:
                cuenta_info = cta
                break

        if cuenta_info:
            percepciones.append(LineaContable(
                cuenta=cuenta_info[0],
                subcuenta=cuenta_info[1],
                concepto=header.strip(),
                monto=monto,
            ))
            continue

        # Buscar en deducciones
        cuenta_info = None
        for concepto, cta in DEDUCCIONES_CUENTAS.items():
            if concepto in header_norm or header_norm in concepto:
                cuenta_info = cta
                break

        if cuenta_info:
            deducciones.append(LineaContable(
                cuenta=cuenta_info[0],
                subcuenta=cuenta_info[1],
                concepto=header.strip(),
                monto=monto,
            ))

    return percepciones, deducciones


def _celda(ws, col: str, fila: int):
    """Obtiene el valor de una celda."""
    return ws[f'{col}{fila}'].value


def _celda_str(ws, col: str, fila: int) -> Optional[str]:
    """Obtiene el valor como string."""
    valor = _celda(ws, col, fila)
    if valor is None:
        return None
    return str(valor).strip() or None


# ---------------------------------------------------------------------------
# Parser de PDF Lista de Raya
# ---------------------------------------------------------------------------

def parsear_lista_raya_pdf(
    ruta: Path,
) -> Tuple[List[LineaContable], List[LineaContable]]:
    """Parsea percepciones y deducciones del PDF Lista de Raya de CONTPAQi.

    El PDF tiene 3 secciones:
      - Percepciones (izquierda) + Deducciones (derecha) en layout 2 columnas
      - Obligaciones patronales (se omiten)
      - Rubros IMSS (se omiten)

    Solo se extraen percepciones y deducciones reales para la poliza contable.
    Items informativos (Subs al Empleo, ISR antes de Subs, Ajuste al neto,
    Infonavit correspondiente) se omiten.

    Args:
        ruta: Ruta al PDF Lista de Raya.

    Returns:
        Tupla (percepciones, deducciones) como listas de LineaContable.
    """
    if pdfplumber is None:
        logger.error("pdfplumber no disponible para parsear Lista de Raya")
        return [], []

    logger.info("Parseando Lista de Raya: {}", ruta.name)

    texto = ''
    with pdfplumber.open(str(ruta)) as pdf:
        for page in pdf.pages:
            texto += page.extract_text() or ''

    if not texto:
        logger.error("No se pudo extraer texto de {}", ruta.name)
        return [], []

    # Encontrar todas las entradas: numero + concepto + monto
    entradas = _RE_ENTRADA_PDF.findall(texto)

    percepciones = []
    deducciones = []

    for num_str, concepto_raw, monto_str in entradas:
        num = int(num_str)
        concepto_raw = concepto_raw.strip()

        # Omitir conceptos informativos y obligaciones patronales
        if num in _CONCEPTOS_OMITIR:
            continue

        # Parsear monto
        try:
            monto = Decimal(monto_str.replace(',', ''))
        except InvalidOperation:
            continue

        if monto <= 0:
            continue

        concepto_norm = _normalizar_texto(concepto_raw)

        # Buscar en percepciones
        cuenta_info = _buscar_cuenta_pdf(concepto_norm, PERCEPCIONES_PDF)
        if cuenta_info:
            percepciones.append(LineaContable(
                concepto=concepto_raw,
                cuenta=cuenta_info[0],
                subcuenta=cuenta_info[1],
                monto=monto,
            ))
            continue

        # Buscar en deducciones
        cuenta_info = _buscar_cuenta_pdf(concepto_norm, DEDUCCIONES_PDF)
        if cuenta_info:
            deducciones.append(LineaContable(
                concepto=concepto_raw,
                cuenta=cuenta_info[0],
                subcuenta=cuenta_info[1],
                monto=monto,
            ))
            continue

        # No reconocido — log para revision
        logger.debug(
            "Lista de Raya: concepto no reconocido #{} '{}' ${:,.2f}",
            num, concepto_raw, monto,
        )

    logger.info(
        "Lista de Raya: {} percepciones (${:,.2f}), {} deducciones (${:,.2f})",
        len(percepciones),
        sum(p.monto for p in percepciones),
        len(deducciones),
        sum(d.monto for d in deducciones),
    )

    return percepciones, deducciones


def _buscar_cuenta_pdf(
    concepto_norm: str, mapeo: dict,
) -> Optional[tuple]:
    """Busca concepto normalizado en mapeo de cuentas contables."""
    # Match exacto
    if concepto_norm in mapeo:
        return mapeo[concepto_norm]
    # Match parcial: clave del mapeo contenida en el concepto
    for key, cuenta in mapeo.items():
        if key in concepto_norm:
            return cuenta
    return None
