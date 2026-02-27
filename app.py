"""Dashboard Streamlit para Conciliacion Bancaria — Agente5.

Interfaz visual para gestion de archivos, ejecucion de conciliacion
y visualizacion de resultados. NO modifica logica existente del Agente5,
solo consume la API de procesar_estado_cuenta().

Uso:
    streamlit run app.py
"""

import io
import json
import os
import shutil
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from loguru import logger

from config.settings import CUENTAS_BANCARIAS, CUENTA_POR_NUMERO, Settings
from src.erp.sav7_connector import SAV7Connector
from src.models import AccionLinea, ResultadoLinea, TipoProceso
from src.orquestador_unificado import procesar_estado_cuenta
from src.reports.reporte_demo import generar_reporte_estado_cuenta

# Ruta al Agente 4 (se importa lazy en _importar_agente4())
_AGENTE4_PATH = Path(__file__).resolve().parent.parent / 'Agente4' / 'conciliacion-sat-erp'

load_dotenv(Path(__file__).resolve().parent / '.env')

# ---------------------------------------------------------------------------
# Configuracion de pagina
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="Conciliacion Bancaria — Agente5",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ---------------------------------------------------------------------------
# Categorias de archivos
# ---------------------------------------------------------------------------

@dataclass
class CategoriaArchivo:
    """Define una categoria de archivo de entrada."""
    id: str
    nombre: str
    extensiones: Tuple[str, ...]
    obligatorio: bool
    multiple: bool
    descripcion: str


CATEGORIAS: List[CategoriaArchivo] = [
    CategoriaArchivo(
        id='01_Estado_de_Cuenta',
        nombre='Estado de Cuenta',
        extensiones=('.xlsx',),
        obligatorio=True,
        multiple=False,
        descripcion='Archivo Excel de BANREGIO con hojas Banregio F, Banregio T, BANREGIO GTS',
    ),
    CategoriaArchivo(
        id='02_Tesoreria',
        nombre='Tesoreria (Ingresos)',
        extensiones=('.xlsx',),
        obligatorio=True,
        multiple=False,
        descripcion='Reporte de ingresos con una hoja por dia (cortes Z, facturas, depositos)',
    ),
    CategoriaArchivo(
        id='03_Nomina',
        nombre='Nomina',
        extensiones=('.xlsx',),
        obligatorio=False,
        multiple=False,
        descripcion='Excel CONTPAQi con hojas NOM, DISPERCION, CHEQUE',
    ),
    CategoriaArchivo(
        id='04_Lista_de_Raya',
        nombre='Lista de Raya',
        extensiones=('.pdf',),
        obligatorio=False,
        multiple=False,
        descripcion='PDF resumen de nomina CONTPAQi',
    ),
    CategoriaArchivo(
        id='05_IMSS',
        nombre='IMSS / INFONAVIT',
        extensiones=('.pdf',),
        obligatorio=False,
        multiple=False,
        descripcion='PDF de resumen de liquidacion IMSS/INFONAVIT',
    ),
    CategoriaArchivo(
        id='06_Impuesto_Federal',
        nombre='Impuesto Federal',
        extensiones=('.pdf',),
        obligatorio=False,
        multiple=True,
        descripcion='PDFs de declaraciones federales (acuses, declaracion completa, detalle IEPS)',
    ),
    CategoriaArchivo(
        id='07_Impuesto_Estatal',
        nombre='Impuesto Estatal',
        extensiones=('.pdf',),
        obligatorio=False,
        multiple=False,
        descripcion='PDF de declaracion 3% sobre nominas',
    ),
]


# ---------------------------------------------------------------------------
# Funciones de infraestructura de archivos
# ---------------------------------------------------------------------------

def obtener_ruta_base() -> Path:
    """Obtiene la ruta base desde variable de entorno."""
    ruta = os.getenv('RUTA_ARCHIVOS', r'\\SERVERMC\Asesoft\ImplementacionIA')
    return Path(ruta)


def obtener_ruta_periodo(periodo: str) -> Path:
    """Retorna la ruta del periodo (ej: 2026-02)."""
    return obtener_ruta_base() / 'data' / periodo


def crear_estructura_periodo(periodo: str) -> Path:
    """Crea todas las carpetas del periodo si no existen."""
    ruta_periodo = obtener_ruta_periodo(periodo)

    for cat in CATEGORIAS:
        (ruta_periodo / cat.id / 'entrada').mkdir(parents=True, exist_ok=True)
        (ruta_periodo / cat.id / 'procesados').mkdir(parents=True, exist_ok=True)

    (ruta_periodo / '09_Reportes').mkdir(parents=True, exist_ok=True)
    (ruta_periodo / 'logs').mkdir(parents=True, exist_ok=True)

    return ruta_periodo


def listar_archivos_entrada(periodo: str, categoria: CategoriaArchivo) -> List[Path]:
    """Lista archivos validos en la carpeta entrada/ de una categoria."""
    carpeta = obtener_ruta_periodo(periodo) / categoria.id / 'entrada'
    if not carpeta.exists():
        return []
    archivos = []
    for ext in categoria.extensiones:
        archivos.extend(carpeta.glob(f'*{ext}'))
        archivos.extend(carpeta.glob(f'*{ext.upper()}'))
    # Deduplicar (en caso de sistemas case-insensitive)
    vistos = set()
    unicos = []
    for a in sorted(archivos, key=lambda p: p.name):
        if a.name.lower() not in vistos:
            vistos.add(a.name.lower())
            unicos.append(a)
    return unicos


def guardar_archivo_subido(periodo: str, categoria: CategoriaArchivo,
                           archivo_subido) -> Path:
    """Guarda un archivo subido via Streamlit en la carpeta entrada/."""
    carpeta = obtener_ruta_periodo(periodo) / categoria.id / 'entrada'
    carpeta.mkdir(parents=True, exist_ok=True)
    # Si la categoria solo acepta 1 archivo, eliminar los existentes
    if not categoria.multiple:
        for existente in listar_archivos_entrada(periodo, categoria):
            existente.unlink()
    destino = carpeta / archivo_subido.name
    with open(destino, 'wb') as f:
        f.write(archivo_subido.getbuffer())
    return destino


def eliminar_archivo(ruta: Path):
    """Elimina un archivo de entrada."""
    if ruta.exists():
        ruta.unlink()


def _buscar_archivo(carpeta: Path, extension: str) -> Optional[Path]:
    """Busca el primer archivo con la extension dada en la carpeta."""
    if not carpeta.exists():
        return None
    archivos = list(carpeta.glob(f'*{extension}'))
    archivos.extend(carpeta.glob(f'*{extension.upper()}'))
    if archivos:
        return sorted(archivos, key=lambda p: p.name)[0]
    return None


# ---------------------------------------------------------------------------
# Ajustes de impuestos (formulario en la app)
# ---------------------------------------------------------------------------

def _ruta_ajustes_json(periodo: str) -> Path:
    """Ruta del JSON de ajustes de impuestos del periodo."""
    return obtener_ruta_periodo(periodo) / 'ajustes_impuestos.json'


def cargar_ajustes(periodo: str) -> Dict:
    """Carga ajustes de impuestos desde JSON del periodo."""
    ruta = _ruta_ajustes_json(periodo)
    if ruta.exists():
        with open(ruta, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def guardar_ajustes(periodo: str, ajustes: Dict):
    """Guarda ajustes de impuestos en JSON del periodo."""
    ruta = _ruta_ajustes_json(periodo)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with open(ruta, 'w', encoding='utf-8') as f:
        json.dump(ajustes, f, indent=2, ensure_ascii=False)


def _generar_excel_ajustes(ajustes: Dict, destino: Path):
    """Genera AJUSTES_IMPUESTOS.xlsx desde los valores del formulario.

    Formato compatible con src/entrada/ajustes_impuestos.py:
    B2=IMSS, B3=IVA Acumulable, B4=IVA Acreditable, filas 7+=retenciones.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    # Headers
    ws.cell(row=1, column=1, value='Concepto')
    ws.cell(row=1, column=2, value='Importe')

    # Montos fijos
    ws.cell(row=2, column=1, value='IMSS')
    if ajustes.get('total_imss'):
        ws.cell(row=2, column=2, value=float(ajustes['total_imss']))

    ws.cell(row=3, column=1, value='IVA Acumulable')
    if ajustes.get('iva_acumulable'):
        ws.cell(row=3, column=2, value=float(ajustes['iva_acumulable']))

    ws.cell(row=4, column=1, value='IVA Acreditable')
    if ajustes.get('iva_acreditable'):
        ws.cell(row=4, column=2, value=float(ajustes['iva_acreditable']))

    # Retenciones IVA
    retenciones = ajustes.get('retenciones_iva', [])
    if retenciones:
        ws.cell(row=6, column=1, value='Proveedor')
        ws.cell(row=6, column=2, value='Nombre')
        ws.cell(row=6, column=3, value='Monto')
        for i, ret in enumerate(retenciones):
            fila = 7 + i
            ws.cell(row=fila, column=1, value=ret.get('proveedor', ''))
            ws.cell(row=fila, column=2, value=ret.get('nombre', ''))
            if ret.get('monto'):
                ws.cell(row=fila, column=3, value=float(ret['monto']))

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))


# ---------------------------------------------------------------------------
# Mapeo de archivos a parametros de la API
# ---------------------------------------------------------------------------

def construir_parametros_api(periodo: str) -> Dict:
    """Mapea archivos en carpetas a parametros de procesar_estado_cuenta()."""
    ruta = obtener_ruta_periodo(periodo)
    params = {}

    # Estado de cuenta (obligatorio)
    ec = _buscar_archivo(ruta / '01_Estado_de_Cuenta' / 'entrada', '.xlsx')
    if ec:
        params['ruta_estado_cuenta'] = ec

    # Tesoreria
    tes = _buscar_archivo(ruta / '02_Tesoreria' / 'entrada', '.xlsx')
    if tes:
        params['ruta_tesoreria'] = tes

    # Nomina
    nom = _buscar_archivo(ruta / '03_Nomina' / 'entrada', '.xlsx')
    if nom:
        params['ruta_nomina'] = nom

    # Lista de raya
    lr = _buscar_archivo(ruta / '04_Lista_de_Raya' / 'entrada', '.pdf')
    if lr:
        params['ruta_lista_raya'] = lr

    # IMSS
    imss = _buscar_archivo(ruta / '05_IMSS' / 'entrada', '.pdf')
    if imss:
        params['ruta_imss'] = imss

    # Impuestos federales (multiples PDFs)
    rutas_imp = {}
    federal_dir = ruta / '06_Impuesto_Federal' / 'entrada'
    if federal_dir.exists():
        pdfs = sorted(federal_dir.glob('*.pdf')) + sorted(federal_dir.glob('*.PDF'))
        for pdf in pdfs:
            nombre = pdf.stem.lower()
            if 'acusepdf' in nombre or 'acuse-1' in nombre:
                rutas_imp['ruta_acuse_federal_1'] = pdf
            elif 'acuse.dcm' in nombre or 'acuse-2' in nombre:
                rutas_imp['ruta_acuse_federal_2'] = pdf
            elif 'declaracion.acuse' in nombre or 'ieps' in nombre:
                rutas_imp['ruta_detalle_ieps'] = pdf
            elif nombre.startswith('dcm') and not 'acuse' in nombre:
                rutas_imp['ruta_declaracion_completa'] = pdf

    # Impuesto estatal
    estatal_dir = ruta / '07_Impuesto_Estatal' / 'entrada'
    if estatal_dir.exists():
        est = _buscar_archivo(estatal_dir, '.pdf')
        if est:
            rutas_imp['ruta_impuesto_estatal'] = est

    if rutas_imp:
        params['rutas_impuestos'] = rutas_imp

    # Ajustes: generar Excel desde formulario para que orquestador lo detecte
    ajustes = cargar_ajustes(periodo)
    if ajustes and 'ruta_estado_cuenta' in params:
        destino = params['ruta_estado_cuenta'].parent / 'AJUSTES_IMPUESTOS.xlsx'
        _generar_excel_ajustes(ajustes, destino)

    return params


# ---------------------------------------------------------------------------
# Conexion a BD (cached)
# ---------------------------------------------------------------------------

@st.cache_resource
def obtener_connector() -> SAV7Connector:
    """Crea y cachea la conexion a BD."""
    settings = Settings.from_env()
    return SAV7Connector(settings)


def verificar_conexion() -> bool:
    """Verifica si la conexion a BD esta activa."""
    try:
        connector = obtener_connector()
        return connector.test_conexion()
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Ejecucion de conciliacion
# ---------------------------------------------------------------------------

def ejecutar_conciliacion(
    periodo: str,
    solo_fecha: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    dry_run: bool = True,
) -> Tuple[List[ResultadoLinea], Optional[Path]]:
    """Ejecuta la conciliacion y retorna resultados + ruta de reporte."""
    params = construir_parametros_api(periodo)

    if 'ruta_estado_cuenta' not in params:
        st.error("No se encontro el archivo de Estado de Cuenta en la carpeta de entrada.")
        return [], None

    connector = obtener_connector()

    resultados = procesar_estado_cuenta(
        **params,
        dry_run=dry_run,
        solo_fecha=solo_fecha,
        fecha_fin=fecha_fin,
        connector=connector,
    )

    # Filtrar por fecha si aplica
    if solo_fecha and not fecha_fin:
        # Dia especifico: incluir pagos del dia anterior
        fecha_ayer = solo_fecha - timedelta(days=1)
        resultados = [
            rl for rl in resultados
            if rl.movimiento.fecha == solo_fecha
            or (
                rl.movimiento.fecha == fecha_ayer
                and rl.tipo_clasificado == TipoProceso.PAGO_PROVEEDOR
                and rl.accion in (AccionLinea.CONCILIAR, AccionLinea.SIN_PROCESAR, AccionLinea.ERROR)
            )
        ]
    elif solo_fecha and fecha_fin:
        # Rango: incluir pagos del dia anterior al inicio
        fecha_ayer = solo_fecha - timedelta(days=1)
        resultados = [
            rl for rl in resultados
            if (solo_fecha <= rl.movimiento.fecha <= fecha_fin)
            or (
                rl.movimiento.fecha == fecha_ayer
                and rl.tipo_clasificado == TipoProceso.PAGO_PROVEEDOR
                and rl.accion in (AccionLinea.CONCILIAR, AccionLinea.SIN_PROCESAR, AccionLinea.ERROR)
            )
        ]

    # Generar reporte si no es dry-run
    ruta_reporte = None
    if not dry_run and resultados:
        ruta_periodo = obtener_ruta_periodo(periodo)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if fecha_fin:
            fecha_str = f'{solo_fecha.strftime("%Y-%m-%d")}_a_{fecha_fin.strftime("%Y-%m-%d")}'
        elif solo_fecha:
            fecha_str = solo_fecha.strftime('%Y-%m-%d')
        else:
            fecha_str = 'completo'
        nombre = f'REPORTE_{fecha_str}_{timestamp}.xlsx'
        ruta_reporte = ruta_periodo / '09_Reportes' / nombre
        ruta_reporte.parent.mkdir(parents=True, exist_ok=True)
        generar_reporte_estado_cuenta(connector, resultados, ruta_reporte)

    # Guardar log de ejecucion
    _guardar_log_ejecucion(periodo, solo_fecha, dry_run, resultados, ruta_reporte)

    return resultados, ruta_reporte


def _guardar_log_ejecucion(
    periodo: str,
    solo_fecha: Optional[date],
    dry_run: bool,
    resultados: List[ResultadoLinea],
    ruta_reporte: Optional[Path],
):
    """Guarda un log JSON de la ejecucion."""
    ruta_periodo = obtener_ruta_periodo(periodo)
    logs_dir = ruta_periodo / 'logs'
    logs_dir.mkdir(parents=True, exist_ok=True)

    from collections import Counter
    acciones = Counter(rl.accion.value for rl in resultados)

    log = {
        'timestamp': datetime.now().isoformat(),
        'periodo': periodo,
        'solo_fecha': solo_fecha.isoformat() if solo_fecha else None,
        'dry_run': dry_run,
        'total_lineas': len(resultados),
        'acciones': dict(acciones),
        'reporte': str(ruta_reporte) if ruta_reporte else None,
    }

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    modo = 'dryrun' if dry_run else 'ejecucion'
    ruta_log = logs_dir / f'{modo}_{timestamp}.json'
    with open(ruta_log, 'w', encoding='utf-8') as f:
        json.dump(log, f, indent=2, ensure_ascii=False)


# ---------------------------------------------------------------------------
# Funciones de visualizacion
# ---------------------------------------------------------------------------

def mostrar_resumen_resultados(resultados: List[ResultadoLinea]):
    """Muestra metricas y tabla de resultados."""
    from collections import Counter

    if not resultados:
        st.warning("No hay resultados para mostrar.")
        return

    acciones = Counter(rl.accion for rl in resultados)
    folios = set()
    for rl in resultados:
        folios.update(rl.folios)

    # Metricas principales
    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Total movimientos", len(resultados))
    col2.metric("INSERT", acciones.get(AccionLinea.INSERT, 0))
    col3.metric("CONCILIAR", acciones.get(AccionLinea.CONCILIAR, 0))
    col4.metric("Folios", len(folios))
    col5.metric("Errores", acciones.get(AccionLinea.ERROR, 0))

    # Tabla detallada
    fecha_max = max(rl.movimiento.fecha for rl in resultados) if resultados else None
    filas = []
    for rl in resultados:
        mov = rl.movimiento
        nota = rl.nota or ''
        if fecha_max and mov.fecha < fecha_max and rl.tipo_clasificado == TipoProceso.PAGO_PROVEEDOR:
            nota = f"[Dia anterior] {nota}" if nota else "[Dia anterior]"
        # Prefijo "Procesado" para acciones exitosas
        if rl.accion in (AccionLinea.INSERT, AccionLinea.CONCILIAR):
            nota = f"Procesado. {nota}" if nota else "Procesado"
        filas.append({
            'Fecha': mov.fecha.strftime('%d/%m/%Y'),
            'Cuenta': CUENTAS_BANCARIAS[CUENTA_POR_NUMERO[mov.cuenta_banco]].nombre if mov.cuenta_banco in CUENTA_POR_NUMERO else mov.cuenta_banco,
            'Descripcion': mov.descripcion[:60],
            'Cargo': float(mov.cargo) if mov.cargo else None,
            'Abono': float(mov.abono) if mov.abono else None,
            'Clasificacion': rl.tipo_clasificado.value,
            'Accion': rl.accion.value,
            'Folios': ', '.join(str(f) for f in rl.folios),
            'Nota': nota,
        })

    df = pd.DataFrame(filas)

    # Colores por accion
    _COLORES_ACCION = {
        'INSERT': 'background-color: #d4edda',       # verde claro
        'CONCILIAR': 'background-color: #cce5ff',    # azul claro
        'OMITIR': 'background-color: #e2e3e5',       # gris claro
        'ERROR': 'background-color: #f8d7da',        # rojo claro
        'SIN_PROCESAR': 'background-color: #fff3cd',  # amarillo claro
        'REQUIERE_REVISION': 'background-color: #fff3cd',
        'DESCONOCIDO': '',
    }

    def _colorear_fila(row):
        estilo = _COLORES_ACCION.get(row['Accion'], '')
        return [estilo] * len(row)

    styled = df.style.apply(_colorear_fila, axis=1)
    st.dataframe(styled, use_container_width=True, height=400)


# ---------------------------------------------------------------------------
# UI: Sidebar
# ---------------------------------------------------------------------------

_PAGINAS = [
    "Agente5 — Conciliacion Bancaria",
    "Agente4 — Conciliacion SAT",
]


def render_sidebar():
    """Renderiza el sidebar con selector de pagina, periodo y estado de conexion."""
    with st.sidebar:
        # Selector de pagina
        pagina = st.radio("Modulo", _PAGINAS, label_visibility="collapsed")
        st.session_state['pagina'] = pagina

        st.divider()

        # Selector de periodo
        hoy = date.today()
        anio = st.number_input("Año", min_value=2024, max_value=2030, value=hoy.year)
        mes = st.number_input("Mes", min_value=1, max_value=12, value=hoy.month)
        periodo = f"{anio}-{mes:02d}"

        st.session_state['periodo'] = periodo

        # Crear estructura automaticamente al seleccionar periodo
        crear_estructura_periodo(periodo)

        st.divider()

        # Estado de conexion
        st.subheader("Base de datos")
        if verificar_conexion():
            st.success("Conectado")
        else:
            st.error("Sin conexion")
            if st.button("Reintentar conexion"):
                st.cache_resource.clear()
                st.rerun()

        st.divider()

        # Info de ruta
        st.caption(f"Ruta base: {obtener_ruta_base()}")
        st.caption(f"Periodo: {periodo}")


# ---------------------------------------------------------------------------
# UI: Tab Archivos
# ---------------------------------------------------------------------------

def render_tab_archivos():
    """Tab de gestion de archivos."""
    periodo = st.session_state.get('periodo', '')
    if not periodo:
        st.info("Seleccione un periodo en el sidebar.")
        return

    st.subheader(f"Archivos del periodo {periodo}")

    for cat in CATEGORIAS:
        archivos = listar_archivos_entrada(periodo, cat)
        tiene_archivo = len(archivos) > 0

        # Icono de estado
        if tiene_archivo:
            icono = "✅"
        elif cat.obligatorio:
            icono = "❌"
        else:
            icono = "⬜"

        with st.expander(f"{icono} {cat.nombre}", expanded=not tiene_archivo and cat.obligatorio):
            st.caption(cat.descripcion)

            # Archivos existentes
            if archivos:
                for arch in archivos:
                    col_nombre, col_accion = st.columns([4, 1])
                    col_nombre.text(f"📄 {arch.name}")
                    if col_accion.button("Eliminar", key=f"del_{cat.id}_{arch.name}"):
                        eliminar_archivo(arch)
                        st.rerun()

            # Uploader
            tipo_ext = cat.extensiones[0].replace('.', '').upper()
            accept = list(cat.extensiones)
            uploaded = st.file_uploader(
                f"Subir {tipo_ext}",
                type=[e.replace('.', '') for e in cat.extensiones],
                accept_multiple_files=cat.multiple,
                key=f"upload_{cat.id}",
            )

            if uploaded:
                archivos_subidos = uploaded if isinstance(uploaded, list) else [uploaded]
                for archivo in archivos_subidos:
                    ruta = guardar_archivo_subido(periodo, cat, archivo)
                    st.success(f"Guardado: {ruta.name}")
                st.rerun()

    # --- Ajustes de impuestos (formulario) ---
    st.divider()
    _render_formulario_ajustes(periodo)

    # Resumen
    st.divider()
    total = len(CATEGORIAS)
    con_archivo = sum(
        1 for cat in CATEGORIAS
        if listar_archivos_entrada(periodo, cat)
    )
    obligatorios_ok = all(
        listar_archivos_entrada(periodo, cat)
        for cat in CATEGORIAS if cat.obligatorio
    )

    col1, col2 = st.columns(2)
    col1.metric("Archivos cargados", f"{con_archivo} / {total}")
    if obligatorios_ok:
        col2.success("Archivos obligatorios completos")
    else:
        col2.warning("Faltan archivos obligatorios")


def _render_formulario_ajustes(periodo: str):
    """Renderiza formulario de ajustes de impuestos."""
    ajustes = cargar_ajustes(periodo)
    tiene_datos = bool(ajustes.get('total_imss') or ajustes.get('iva_acumulable')
                       or ajustes.get('iva_acreditable') or ajustes.get('retenciones_iva'))
    icono = "✅" if tiene_datos else "⬜"

    with st.expander(f"{icono} Ajustes de Impuestos (opcional)", expanded=False):
        st.caption(
            "Override manual de valores extraidos de los PDFs. "
            "Deje en blanco (0) para usar el valor del PDF."
        )

        col1, col2, col3 = st.columns(3)
        with col1:
            total_imss = st.number_input(
                "IMSS",
                min_value=0.0,
                value=float(ajustes.get('total_imss', 0)),
                step=0.01,
                format="%.2f",
                key="ajuste_imss",
                help="Monto total IMSS/INFONAVIT",
            )
        with col2:
            iva_acumulable = st.number_input(
                "IVA Acumulable",
                min_value=0.0,
                value=float(ajustes.get('iva_acumulable', 0)),
                step=0.01,
                format="%.2f",
                key="ajuste_iva_acum",
                help="IVA acumulable de declaracion federal",
            )
        with col3:
            iva_acreditable = st.number_input(
                "IVA Acreditable",
                min_value=0.0,
                value=float(ajustes.get('iva_acreditable', 0)),
                step=0.01,
                format="%.2f",
                key="ajuste_iva_acred",
                help="IVA acreditable de declaracion federal",
            )

        # Retenciones IVA
        st.markdown("**Retenciones IVA por proveedor**")
        retenciones = ajustes.get('retenciones_iva', [])

        # Inicializar session_state para retenciones
        if 'retenciones_lista' not in st.session_state:
            st.session_state['retenciones_lista'] = retenciones if retenciones else []

        retenciones_editadas = []
        for i, ret in enumerate(st.session_state['retenciones_lista']):
            c1, c2, c3, c4 = st.columns([2, 3, 2, 1])
            with c1:
                prov = st.text_input(
                    "Clave", value=ret.get('proveedor', ''),
                    key=f"ret_prov_{i}", label_visibility="collapsed",
                    placeholder="Clave proveedor",
                )
            with c2:
                nombre = st.text_input(
                    "Nombre", value=ret.get('nombre', ''),
                    key=f"ret_nom_{i}", label_visibility="collapsed",
                    placeholder="Nombre",
                )
            with c3:
                monto = st.number_input(
                    "Monto", value=float(ret.get('monto', 0)),
                    min_value=0.0, step=0.01, format="%.2f",
                    key=f"ret_monto_{i}", label_visibility="collapsed",
                )
            with c4:
                if st.button("🗑️", key=f"ret_del_{i}"):
                    st.session_state['retenciones_lista'].pop(i)
                    st.rerun()
            retenciones_editadas.append({
                'proveedor': prov, 'nombre': nombre, 'monto': monto,
            })

        if st.button("+ Agregar retencion", key="agregar_retencion"):
            st.session_state['retenciones_lista'].append(
                {'proveedor': '', 'nombre': '', 'monto': 0}
            )
            st.rerun()

        # Boton guardar
        if st.button("Guardar ajustes", type="primary", key="guardar_ajustes"):
            nuevos_ajustes = {}
            if total_imss > 0:
                nuevos_ajustes['total_imss'] = total_imss
            if iva_acumulable > 0:
                nuevos_ajustes['iva_acumulable'] = iva_acumulable
            if iva_acreditable > 0:
                nuevos_ajustes['iva_acreditable'] = iva_acreditable
            # Filtrar retenciones con datos
            rets_validas = [
                r for r in retenciones_editadas
                if r.get('proveedor') and r.get('monto', 0) > 0
            ]
            if rets_validas:
                nuevos_ajustes['retenciones_iva'] = rets_validas

            guardar_ajustes(periodo, nuevos_ajustes)
            st.session_state['retenciones_lista'] = rets_validas
            st.success("Ajustes guardados")


# ---------------------------------------------------------------------------
# UI: Tab Procesar
# ---------------------------------------------------------------------------

def render_tab_procesar():
    """Tab de ejecucion de conciliacion."""
    periodo = st.session_state.get('periodo', '')
    if not periodo:
        st.info("Seleccione un periodo en el sidebar.")
        return

    # Verificar archivos obligatorios
    params = construir_parametros_api(periodo)
    if 'ruta_estado_cuenta' not in params:
        st.warning(
            "No se encontro el Estado de Cuenta. "
            "Suba el archivo en la pestana Archivos."
        )
        return

    st.subheader("Archivos detectados")
    nombres_params = {
        'ruta_estado_cuenta': 'Estado de Cuenta',
        'ruta_tesoreria': 'Tesoreria',
        'ruta_nomina': 'Nomina',
        'ruta_lista_raya': 'Lista de Raya',
        'ruta_imss': 'IMSS',
        'rutas_impuestos': 'Impuestos',
    }
    for clave, nombre in nombres_params.items():
        if clave in params:
            valor = params[clave]
            if isinstance(valor, dict):
                st.text(f"  ✅ {nombre}: {len(valor)} archivos")
            else:
                st.text(f"  ✅ {nombre}: {valor.name}")
        else:
            st.text(f"  ⬜ {nombre}: no proporcionado")

    st.divider()

    # Selector de fecha
    st.subheader("Configuracion")
    partes = periodo.split('-')
    anio, mes = int(partes[0]), int(partes[1])

    modo_fecha = st.radio(
        "Modo",
        ["Dia especifico", "Rango de fechas"],
        horizontal=True,
    )

    solo_fecha = None
    fecha_fin = None

    if modo_fecha == "Dia especifico":
        dia = st.number_input("Dia", min_value=1, max_value=31, value=1)
        try:
            solo_fecha = date(anio, mes, dia)
        except ValueError:
            st.error("Fecha invalida")
            return
    else:
        col_inicio, col_fin = st.columns(2)
        with col_inicio:
            dia_inicio = st.number_input("Dia inicio", min_value=1, max_value=31, value=1)
        with col_fin:
            dia_fin = st.number_input("Dia fin", min_value=1, max_value=31, value=min(7, 28))
        try:
            solo_fecha = date(anio, mes, dia_inicio)
            fecha_fin = date(anio, mes, dia_fin)
        except ValueError:
            st.error("Fecha invalida")
            return
        if fecha_fin < solo_fecha:
            st.error("La fecha fin debe ser mayor o igual a la fecha inicio.")
            return
        dias_rango = (fecha_fin - solo_fecha).days + 1
        if dias_rango > 7:
            st.error("El rango maximo es de 7 dias.")
            return
        st.caption(f"Rango: {solo_fecha.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')} ({dias_rango} dias)")

    st.divider()

    # Botones de ejecucion
    col_preview, col_ejecutar = st.columns(2)

    with col_preview:
        if st.button("Vista previa (dry-run)", type="secondary", use_container_width=True):
            with st.spinner("Ejecutando vista previa..."):
                resultados, _ = ejecutar_conciliacion(
                    periodo, solo_fecha=solo_fecha, fecha_fin=fecha_fin, dry_run=True,
                )
            st.session_state['ultimos_resultados'] = resultados
            st.session_state['ultimo_modo'] = 'dry-run'

    with col_ejecutar:
        if st.button("Ejecutar conciliacion", type="primary", use_container_width=True):
            if not verificar_conexion():
                st.error("No hay conexion a la base de datos.")
                return
            with st.spinner("Ejecutando conciliacion..."):
                resultados, ruta_reporte = ejecutar_conciliacion(
                    periodo, solo_fecha=solo_fecha, fecha_fin=fecha_fin, dry_run=False,
                )
            st.session_state['ultimos_resultados'] = resultados
            st.session_state['ultimo_modo'] = 'ejecucion'
            if ruta_reporte:
                st.session_state['ultimo_reporte'] = str(ruta_reporte)

    # Mostrar resultados
    if 'ultimos_resultados' in st.session_state:
        modo = st.session_state.get('ultimo_modo', '')
        st.divider()
        st.subheader(f"Resultados ({modo})")
        mostrar_resumen_resultados(st.session_state['ultimos_resultados'])

        # Descargar reporte
        if 'ultimo_reporte' in st.session_state and modo == 'ejecucion':
            ruta_rep = Path(st.session_state['ultimo_reporte'])
            if ruta_rep.exists():
                with open(ruta_rep, 'rb') as f:
                    st.download_button(
                        "Descargar reporte Excel",
                        data=f.read(),
                        file_name=ruta_rep.name,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    )


# ---------------------------------------------------------------------------
# UI: Tab Historial
# ---------------------------------------------------------------------------

def render_tab_historial():
    """Tab de historial de ejecuciones."""
    periodo = st.session_state.get('periodo', '')
    if not periodo:
        st.info("Seleccione un periodo en el sidebar.")
        return

    ruta_periodo = obtener_ruta_periodo(periodo)

    # Logs
    st.subheader("Ejecuciones")
    logs_dir = ruta_periodo / 'logs'
    if logs_dir.exists():
        logs = sorted(logs_dir.glob('*.json'), reverse=True)
        if logs:
            for log_file in logs[:20]:  # Ultimos 20
                try:
                    with open(log_file, 'r', encoding='utf-8') as f:
                        log = json.load(f)
                    modo = "DRY-RUN" if log.get('dry_run') else "EJECUCION"
                    ts = log.get('timestamp', '')[:19]
                    fecha = log.get('solo_fecha', 'completo')
                    total = log.get('total_lineas', 0)
                    acciones = log.get('acciones', {})

                    with st.expander(f"{modo} | {ts} | Fecha: {fecha} | {total} lineas"):
                        for accion, conteo in acciones.items():
                            st.text(f"  {accion}: {conteo}")
                        if log.get('reporte'):
                            st.text(f"  Reporte: {log['reporte']}")
                except Exception:
                    continue
        else:
            st.info("No hay ejecuciones registradas.")
    else:
        st.info("No hay ejecuciones registradas.")

    # Reportes
    st.divider()
    st.subheader("Reportes generados")
    reportes_dir = ruta_periodo / '09_Reportes'
    if reportes_dir.exists():
        reportes = sorted(reportes_dir.glob('*.xlsx'), reverse=True)
        if reportes:
            for rep in reportes[:10]:
                col_nombre, col_descarga = st.columns([3, 1])
                col_nombre.text(f"📊 {rep.name}")
                with open(rep, 'rb') as f:
                    col_descarga.download_button(
                        "Descargar",
                        data=f.read(),
                        file_name=rep.name,
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        key=f"dl_{rep.name}",
                    )
        else:
            st.info("No hay reportes generados.")
    else:
        st.info("No hay reportes generados.")


# ---------------------------------------------------------------------------
# UI: Tab Conciliacion SAT
# ---------------------------------------------------------------------------


def _importar_agente4():
    """Importa modulos del Agente 4 (lazy). Retorna None si no disponible.

    Usa manipulacion temporal de sys.path para que 'from src.xxx' resuelva
    al Agente 4 y no al Agente 5 (ambos tienen carpeta src/).
    """
    if not _AGENTE4_PATH.exists():
        return None
    a4 = str(_AGENTE4_PATH)
    try:
        sys.path.insert(0, a4)
        # Limpiar cache de modulos src.* del Agente 5 temporalmente
        modulos_a5 = {k: v for k, v in sys.modules.items()
                      if k.startswith('src.') or k == 'src'}
        for k in modulos_a5:
            del sys.modules[k]

        from src.sat.descarga import DescargadorSAT
        from src.sat.parser import ParserCFDI
        from src.erp.sav7 import ConectorSAV7 as ConectorSAV7_A4
        from src.conciliacion.comparador import Comparador, ResultadoConciliacion
        from src.reportes.generador import GeneradorReportes

        return {
            'DescargadorSAT': DescargadorSAT,
            'ParserCFDI': ParserCFDI,
            'ConectorSAV7': ConectorSAV7_A4,
            'Comparador': Comparador,
            'ResultadoConciliacion': ResultadoConciliacion,
            'GeneradorReportes': GeneradorReportes,
        }
    except ImportError as e:
        logger.error("No se pudo importar Agente 4: {}", e)
        return None
    finally:
        # Restaurar: quitar Agente 4 del path y re-importar modulos Agente 5
        if a4 in sys.path:
            sys.path.remove(a4)
        # Limpiar modulos del Agente 4 para que src.* vuelva a resolver al Agente 5
        for k in list(sys.modules.keys()):
            if k.startswith('src.') or k == 'src':
                del sys.modules[k]
        # Restaurar modulos Agente 5
        sys.modules.update(modulos_a5)


def _ejecutar_conciliacion_sat(fecha_inicio: date, fecha_fin: date, dir_xmls: Path, dir_reportes: Path):
    """Ejecuta el pipeline completo de conciliacion SAT-ERP via Agente 4."""
    modulos = _importar_agente4()
    if not modulos:
        st.error("No se pudieron importar los modulos del Agente 4.")
        return None

    rfc = os.getenv('SAT_RFC', '')
    cert_path = os.getenv('SAT_CERT_PATH', '')
    key_path = os.getenv('SAT_KEY_PATH', '')
    key_password = os.getenv('SAT_KEY_PASSWORD', '').strip()

    # Mapear env vars de Agente 5 → params de Agente 4
    db_host = os.getenv('DB_SERVER', 'localhost')
    db_name = os.getenv('DB_DATABASE', 'DBSAV71A')
    db_user = os.getenv('DB_USERNAME', '')
    db_password = os.getenv('DB_PASSWORD', '')

    dir_xmls.mkdir(parents=True, exist_ok=True)
    dir_reportes.mkdir(parents=True, exist_ok=True)

    fecha_inicio_str = fecha_inicio.isoformat()
    fecha_fin_str = fecha_fin.isoformat()

    DescargadorSAT = modulos['DescargadorSAT']
    ParserCFDI = modulos['ParserCFDI']
    ConectorSAV7_A4 = modulos['ConectorSAV7']
    Comparador = modulos['Comparador']
    GeneradorReportes = modulos['GeneradorReportes']

    # Pre-cargar certificados en nuestro contexto
    # (evita problemas de import context del Agente 4)
    import base64 as _b64
    from cryptography import x509 as _x509
    from cryptography.hazmat.primitives import serialization as _ser
    from cryptography.hazmat.backends import default_backend as _backend

    with open(cert_path, 'rb') as f:
        _cert_data = f.read()
    try:
        _certificado = _x509.load_pem_x509_certificate(_cert_data, _backend())
    except Exception:
        _certificado = _x509.load_der_x509_certificate(_cert_data, _backend())
    _cert_der = _certificado.public_bytes(_ser.Encoding.DER)
    _cert_b64 = _b64.b64encode(_cert_der).decode()

    with open(key_path, 'rb') as f:
        _key_data = f.read()
    _llave = _ser.load_pem_private_key(_key_data, password=None, backend=_backend())

    with st.status("Conciliacion SAT-ERP en progreso...", expanded=True) as status:
        # Paso 1: Descargar CFDIs del SAT
        st.write("Descargando CFDIs del SAT (puede tardar varios minutos)...")

        # Crear DescargadorSAT sin que cargue certs (usamos los nuestros)
        DescargadorSAT._cargar_certificados_original = DescargadorSAT._cargar_certificados
        DescargadorSAT._cargar_certificados = lambda self: None  # Skip
        try:
            descargador = DescargadorSAT(
                rfc=rfc,
                cert_path=cert_path,
                key_path=key_path,
                key_password=key_password,
                output_dir=str(dir_xmls),
            )
        finally:
            DescargadorSAT._cargar_certificados = DescargadorSAT._cargar_certificados_original
            del DescargadorSAT._cargar_certificados_original

        # Inyectar certs pre-cargados
        descargador.certificado = _certificado
        descargador.cert_der = _cert_der
        descargador.cert_base64 = _cert_b64
        descargador.llave_privada = _llave
        uuids_sat = descargador.descargar_cfdis(
            fecha_inicio=fecha_inicio_str,
            fecha_fin=fecha_fin_str,
            tipo_comprobante='I',
            tipo_descarga='recibidos',
        )
        st.write(f"Descargados: {len(uuids_sat)} CFDIs del SAT")

        # Paso 2: Parsear datos completos de XMLs
        st.write("Parseando datos de XMLs...")
        parser = ParserCFDI(str(dir_xmls))
        datos_sat = parser.procesar_directorio_completo(tipo_filtro=None)
        st.write(f"Parseados: {len(datos_sat)} CFDIs con datos completos")

        # Paso 3: Consultar ERP
        st.write("Consultando ERP...")
        with ConectorSAV7_A4(
            host=db_host,
            database=db_name,
            user=db_user,
            password=db_password,
        ) as erp:
            uuids_erp_periodo = erp.obtener_uuids_por_fechas(fecha_inicio_str, fecha_fin_str)
            uuids_erp_todos = erp.obtener_todos_los_uuids()
            datos_erp = erp.obtener_datos_por_uuid()
        st.write(f"ERP: {len(uuids_erp_periodo)} en periodo, {len(uuids_erp_todos)} totales")

        # Paso 4: Conciliar
        st.write("Conciliando...")
        comparador = Comparador()
        resultado = comparador.conciliar(uuids_sat, uuids_erp_periodo, uuids_erp_todos)
        resultado.fecha_inicio = fecha_inicio_str
        resultado.fecha_fin = fecha_fin_str
        resultado.datos_sat = datos_sat
        resultado.datos_erp = datos_erp

        # Paso 5: Generar reporte CSV
        st.write("Generando reporte...")
        generador = GeneradorReportes(str(dir_reportes))
        ruta_csv = generador.generar_csv_completo(resultado)
        st.write(f"Reporte generado: {Path(ruta_csv).name}")

        status.update(label="Conciliacion completada", state="complete", expanded=False)

    return resultado, ruta_csv


def render_tab_sat():
    """Tab de Conciliacion SAT-ERP (Agente 4)."""
    periodo = st.session_state.get('periodo', '')
    if not periodo:
        st.info("Seleccione un periodo en el sidebar.")
        return

    st.header("Conciliacion SAT vs ERP")
    st.caption("Descarga CFDIs del SAT y compara UUIDs contra el ERP (Agente 4)")

    # Verificar disponibilidad del Agente 4
    if not _AGENTE4_PATH.exists():
        st.error(f"No se encontro el Agente 4 en: {_AGENTE4_PATH}")
        return

    # Verificar credenciales FIEL
    rfc = os.getenv('SAT_RFC', '')
    cert_path = os.getenv('SAT_CERT_PATH', '')
    key_path = os.getenv('SAT_KEY_PATH', '')

    fiel_ok = True
    if not rfc:
        st.warning("Variable SAT_RFC no configurada en .env")
        fiel_ok = False
    if not cert_path or not Path(cert_path).exists():
        st.warning(f"Certificado FIEL no encontrado: {cert_path or '(no configurado)'}")
        fiel_ok = False
    if not key_path or not Path(key_path).exists():
        st.warning(f"Llave privada FIEL no encontrada: {key_path or '(no configurado)'}")
        fiel_ok = False

    if not fiel_ok:
        st.error("Configure las credenciales FIEL en el archivo .env para continuar.")
        return

    st.success(f"FIEL configurada — RFC: {rfc}")

    # Selector de fechas
    anio, mes = periodo.split('-')
    anio, mes = int(anio), int(mes)
    primer_dia = date(anio, mes, 1)
    if mes == 12:
        ultimo_dia = date(anio + 1, 1, 1) - timedelta(days=1)
    else:
        ultimo_dia = date(anio, mes + 1, 1) - timedelta(days=1)

    col1, col2 = st.columns(2)
    with col1:
        fecha_inicio = st.date_input("Fecha inicio", value=primer_dia, key="sat_fecha_inicio")
    with col2:
        fecha_fin = st.date_input("Fecha fin", value=ultimo_dia, key="sat_fecha_fin")

    if fecha_inicio > fecha_fin:
        st.error("La fecha inicio debe ser anterior a la fecha fin.")
        return

    # Directorios de salida
    ruta_periodo = obtener_ruta_periodo(periodo)
    dir_xmls = ruta_periodo / '10_SAT' / 'xmls'
    dir_reportes = ruta_periodo / '10_SAT' / 'reportes'

    # Boton ejecutar
    if st.button("Ejecutar conciliacion SAT", type="primary", key="btn_sat"):
        try:
            result = _ejecutar_conciliacion_sat(fecha_inicio, fecha_fin, dir_xmls, dir_reportes)
            if result:
                resultado, ruta_csv = result
                st.session_state['sat_resultado'] = resultado
                st.session_state['sat_csv_path'] = ruta_csv
        except Exception as e:
            st.error(f"Error durante la conciliacion: {e}")
            logger.exception("Error en conciliacion SAT")

    # Mostrar resultados si existen
    resultado = st.session_state.get('sat_resultado')
    ruta_csv = st.session_state.get('sat_csv_path')

    if resultado:
        st.divider()

        # Metricas
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Conciliados", resultado.total_conciliados)
        col2.metric("Faltantes en ERP", resultado.total_faltantes_erp)
        col3.metric("Faltantes en SAT", resultado.total_faltantes_sat)
        pct = resultado.porcentaje_conciliacion
        col4.metric("Conciliacion", f"{pct:.1f}%")

        # Faltantes en ERP
        if resultado.faltantes_en_erp:
            with st.expander(f"Faltantes en ERP ({resultado.total_faltantes_erp})", expanded=True):
                filas = []
                for uuid in sorted(resultado.faltantes_en_erp):
                    datos = resultado.datos_sat.get(uuid, {})
                    filas.append({
                        'UUID': uuid[:8] + '...',
                        'RFC Emisor': datos.get('rfc_emisor', ''),
                        'Emisor': (datos.get('nombre_emisor', '') or '')[:40],
                        'Total': f"${float(datos.get('total', 0)):,.2f}",
                        'Fecha': datos.get('fecha', ''),
                    })
                st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

        # Faltantes en SAT
        if resultado.faltantes_en_sat:
            with st.expander(f"Faltantes en SAT ({resultado.total_faltantes_sat})"):
                filas = []
                for uuid in sorted(resultado.faltantes_en_sat):
                    datos = resultado.datos_erp.get(uuid, {})
                    filas.append({
                        'UUID': uuid[:8] + '...',
                        'Serie': datos.get('serie', ''),
                        'NumRec': datos.get('numrec', ''),
                        'Proveedor': (datos.get('proveedor_nombre', '') or '')[:40],
                        'Total': f"${float(datos.get('total', 0)):,.2f}",
                    })
                st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

        # Boton descarga CSV
        if ruta_csv and Path(ruta_csv).exists():
            st.divider()
            with open(ruta_csv, 'rb') as f:
                st.download_button(
                    "Descargar reporte CSV completo",
                    data=f.read(),
                    file_name=Path(ruta_csv).name,
                    mime='text/csv',
                )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    render_sidebar()

    pagina = st.session_state.get('pagina', _PAGINAS[0])

    if pagina == _PAGINAS[0]:
        # Agente 5 — Conciliacion Bancaria
        tab_archivos, tab_procesar, tab_historial = st.tabs([
            "📁 Archivos",
            "⚙️ Procesar",
            "📋 Historial",
        ])

        with tab_archivos:
            render_tab_archivos()

        with tab_procesar:
            render_tab_procesar()

        with tab_historial:
            render_tab_historial()

    else:
        # Agente 4 — Conciliacion SAT
        render_tab_sat()


if __name__ == '__main__':
    main()
